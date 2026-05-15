import argparse
import importlib
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from formatters import formatter_base


SERVICE_ORDER = [
    "compute",
    "instance_pools",
    "vcn",
    "waf",
    "waf_edge",
    "mysql",
    "dbcs",
    "adb",
    "vpn",
    "fastconnect",
    "file_storage",
    "block_storage",
    "object_storage",
    "load_balancer",
    "network_load_balancer",
    "dns",
]

COMPUTE_REQUIRED_CONTAINERS = {
    "compute_raw",
    "networking_enriched",
    "storage_enriched",
    "_errors",
}

COMPUTE_REQUIRED_NETWORKING_KEYS = [
    "public_ip",
    "private_ip",
    "vcn_name",
    "subnet_name",
]

COMPUTE_REQUIRED_STORAGE_KEYS = [
    "boot_volume_name",
    "boot_volume_size_in_gbs",
    "block_volume_name",
    "block_volume_size_in_gbs",
    "block_volume_attachment_type",
]


class CheckState:
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.ok_count = 0
        self.warnings = []
        self.failures = []

    def ok(self, message):
        self.ok_count += 1
        if self.verbose:
            print(f"[OK] {message}")

    def warn(self, message):
        self.warnings.append(message)
        print(f"[WARN] {message}")

    def fail(self, message):
        self.failures.append(message)
        print(f"[FAIL] {message}")


def _read_json(path, state):
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        state.fail(f"{path}: JSON read failed: {exc}")
        return None


def _service_from_raw_path(path, profile, state):
    suffix = f"_{profile}.json"
    if not path.name.endswith(suffix):
        state.fail(
            f"{path}: raw file name must follow raw_data/<profile>/<service>_<profile>.json"
        )
        return None
    return path.name[: -len(suffix)]


def _discover_profiles(raw_root):
    if not raw_root.exists():
        return []
    return sorted(p.name for p in raw_root.iterdir() if p.is_dir())


def _ordered_services(services):
    order = {name: idx for idx, name in enumerate(SERVICE_ORDER)}
    return sorted(services, key=lambda service: (order.get(service, len(order)), service))


def _raw_paths_for_profile(raw_root, profile, services, state):
    profile_dir = raw_root / profile
    if not profile_dir.is_dir():
        state.fail(f"{profile_dir}: raw profile directory is missing")
        return {}

    paths = {}
    if services:
        for service in services:
            paths[service] = profile_dir / f"{service}_{profile}.json"
    else:
        for path in sorted(profile_dir.glob("*.json")):
            service = _service_from_raw_path(path, profile, state)
            if service:
                paths[service] = path

    if not paths:
        state.fail(f"{profile_dir}: no raw JSON files found")
    return {service: paths[service] for service in _ordered_services(paths)}


def _check_top_level_schema(path, service, data, state):
    if not isinstance(data, list):
        state.fail(f"{path}: top-level JSON must be list[dict]")
        return
    state.ok(f"{path}: JSON top-level is a list")

    if not data:
        state.warn(f"{path}: raw data is an empty list")
        return

    expected_raw_container = f"{service}_raw"
    for idx, row in enumerate(data, 1):
        prefix = f"{path} row={idx}"
        if not isinstance(row, dict):
            state.fail(f"{prefix}: row must be a dict")
            continue

        if "_errors" not in row:
            state.fail(f"{prefix}: missing _errors")
        elif not isinstance(row["_errors"], list):
            state.fail(f"{prefix}: _errors must be a list")

        if expected_raw_container not in row:
            state.fail(f"{prefix}: missing {expected_raw_container}")

        container_keys = [key for key in row if key != "_errors"]
        if not container_keys:
            state.fail(f"{prefix}: no domain containers found")
        for key in container_keys:
            if not isinstance(row[key], dict):
                state.fail(f"{prefix}: top-level key {key} must be a dict container")

    state.ok(f"{path}: checked {len(data)} raw rows")


def _check_compute_contract(path, data, state):
    if not data:
        return

    for idx, row in enumerate(data, 1):
        prefix = f"{path} row={idx}"
        missing = COMPUTE_REQUIRED_CONTAINERS.difference(row)
        if missing:
            state.fail(f"{prefix}: missing compute containers {sorted(missing)}")
            continue

        networking = row.get("networking_enriched")
        storage = row.get("storage_enriched")
        if not isinstance(networking, dict) or not isinstance(storage, dict):
            state.fail(f"{prefix}: compute enrichment containers must be dicts")
            continue

        for key in COMPUTE_REQUIRED_NETWORKING_KEYS:
            if key not in networking:
                state.fail(f"{prefix}: networking_enriched missing {key}")
        for key in COMPUTE_REQUIRED_STORAGE_KEYS:
            if key not in storage:
                state.fail(f"{prefix}: storage_enriched missing {key}")

    state.ok(f"{path}: checked compute contract")


def _display_column_name(column):
    return formatter_base._to_display_column_name(column)


def _build_sheet_name(sheet_name, used_sheet_names):
    canonical_sheet_name = formatter_base._canonical_sheet_name(str(sheet_name))
    _, group_no = formatter_base._get_group_info(str(sheet_name))
    prefixed_sheet_name = f"{group_no}-{canonical_sheet_name}"
    valid_sheet_name = str(prefixed_sheet_name)[:31]

    if valid_sheet_name in used_sheet_names:
        seq = 2
        while True:
            suffix = f"_{seq}"
            trimmed = str(prefixed_sheet_name)[: max(1, 31 - len(suffix))]
            candidate = f"{trimmed}{suffix}"
            if candidate not in used_sheet_names:
                valid_sheet_name = candidate
                break
            seq += 1

    used_sheet_names.add(valid_sheet_name)
    return valid_sheet_name


def _load_formatter(service, state):
    try:
        return importlib.import_module(f"formatters.formatter_{service}")
    except ImportError:
        state.warn(f"service={service}: specialized formatter not found")
        return None


def _expected_report_sheets(raw_paths, raw_cache, state):
    expected = {}
    used_sheet_names = set()

    for service, path in raw_paths.items():
        data = raw_cache.get(service)
        if data is None:
            continue

        df = pd.json_normalize(data)
        df.insert(0, "category", service.capitalize())

        module = _load_formatter(service, state)
        if module is None:
            processed_sheets = {service.capitalize(): df}
            pref_cols_map = {}
        else:
            try:
                transformed = module.transform(df)
                processed_sheets = (
                    transformed if isinstance(transformed, dict) else {service.capitalize(): transformed}
                )
                preferred = module.get_preferred_columns()
                if isinstance(preferred, dict):
                    pref_cols_map = preferred
                else:
                    first_sheet = list(processed_sheets.keys())[0]
                    pref_cols_map = {first_sheet: preferred}
            except Exception as exc:
                state.fail(f"service={service}: formatter transform failed: {exc}")
                continue

        for sheet_name, sheet_df in processed_sheets.items():
            if not isinstance(sheet_df, pd.DataFrame):
                state.fail(f"service={service} sheet={sheet_name}: formatter did not return a DataFrame")
                continue

            sheet_df = sheet_df.copy()
            preferred_columns = pref_cols_map.get(sheet_name, [])
            existing_pref = []

            if preferred_columns:
                sheet_df, resolved_pref = formatter_base._align_preferred_with_raw(
                    sheet_df,
                    preferred_columns,
                )
                existing_pref = [c for c in resolved_pref if c in sheet_df.columns]
                remaining = [c for c in sheet_df.columns if c not in existing_pref]
                sheet_df = sheet_df[existing_pref + remaining]

            valid_sheet_name = _build_sheet_name(sheet_name, used_sheet_names)
            if "category" in sheet_df.columns:
                sheet_df = sheet_df.drop(columns=["category"])
            sheet_df.insert(
                0,
                "category",
                formatter_base._category_from_prefixed_sheet_name(valid_sheet_name),
            )
            sheet_df = formatter_base._reorder_common_context_columns(sheet_df)

            preferred_names = {"category", *existing_pref}
            preferred_positions = [
                idx for idx, col in enumerate(sheet_df.columns) if col in preferred_names
            ]
            display_headers = [_display_column_name(col) for col in sheet_df.columns]
            last_preferred = max(preferred_positions) if preferred_positions else 0

            expected[valid_sheet_name] = {
                "service": service,
                "sheet": sheet_name,
                "headers": display_headers,
                "preferred_front": display_headers[: last_preferred + 1],
            }

    return expected


def _worksheet_headers(ws):
    if ws.max_row < 1:
        return []
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _check_report(profile, report_dir, expected_sheets, state):
    report_path = report_dir / f"OCI_Report_{profile}.xlsx"
    if not report_path.exists():
        state.fail(f"{report_path}: report file is missing")
        return

    wb = load_workbook(report_path, read_only=True, data_only=True)
    if not wb.sheetnames or wb.sheetnames[0] != "Summary":
        state.fail(f"{report_path}: Summary sheet must be first")
    else:
        state.ok(f"{report_path}: Summary sheet is first")

    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        if summary["A1"].value != "OCI Resource Report":
            state.fail(f"{report_path}: Summary A1 title mismatch")
        else:
            state.ok(f"{report_path}: Summary title is valid")

    for title in wb.sheetnames:
        if title == "Summary":
            continue

        ws = wb[title]
        headers = _worksheet_headers(ws)
        if not headers or headers[0] != "category":
            state.fail(f"{report_path} sheet={title}: first column must be category")
            continue

        expected_category = title.split("-", 1)[1] if "-" in title else title
        for row_idx, (category_value,) in enumerate(
            ws.iter_rows(min_row=2, max_col=1, values_only=True),
            2,
        ):
            if category_value != expected_category:
                state.fail(
                    f"{report_path} sheet={title} row={row_idx}: category must be {expected_category}"
                )
                break
        else:
            state.ok(f"{report_path} sheet={title}: category values are valid")

    for title, expected in expected_sheets.items():
        if title not in wb.sheetnames:
            state.fail(
                f"{report_path}: expected sheet {title} is missing "
                f"(service={expected['service']} sheet={expected['sheet']})"
            )
            continue

        actual_headers = _worksheet_headers(wb[title])
        expected_front = expected["preferred_front"]
        actual_front = actual_headers[: len(expected_front)]
        if actual_front != expected_front:
            state.fail(
                f"{report_path} sheet={title}: preferred column front mismatch "
                f"expected={expected_front} actual={actual_front}"
            )
        else:
            state.ok(f"{report_path} sheet={title}: preferred column front is valid")


def verify_profile(profile, raw_root, report_dir, services, skip_report, state):
    raw_paths = _raw_paths_for_profile(raw_root, profile, services, state)
    raw_cache = {}

    for service, path in raw_paths.items():
        if not path.exists():
            state.fail(f"{path}: raw file is missing")
            continue

        data = _read_json(path, state)
        if data is None:
            continue

        raw_cache[service] = data
        _check_top_level_schema(path, service, data, state)
        if service == "compute":
            _check_compute_contract(path, data, state)

    if not skip_report:
        expected_sheets = _expected_report_sheets(raw_paths, raw_cache, state)
        _check_report(profile, report_dir, expected_sheets, state)


def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Verify raw_data and Excel report contracts for OCI ResourceExtractor."
    )
    parser.add_argument(
        "profiles",
        nargs="*",
        help="Profile name(s) under raw_data/. Defaults to all discovered profiles.",
    )
    parser.add_argument(
        "--service",
        action="append",
        dest="services",
        help="Limit verification to one service. Can be repeated.",
    )
    parser.add_argument("--raw-root", default="raw_data")
    parser.add_argument("--report-dir", default="OCI_Reports")
    parser.add_argument("--skip-report", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv or sys.argv[1:])
    raw_root = Path(args.raw_root)
    report_dir = Path(args.report_dir)
    state = CheckState(verbose=args.verbose)

    profiles = args.profiles or _discover_profiles(raw_root)
    if not profiles:
        state.fail(f"{raw_root}: no profiles found")
    for profile in profiles:
        verify_profile(
            profile=profile,
            raw_root=raw_root,
            report_dir=report_dir,
            services=args.services,
            skip_report=args.skip_report,
            state=state,
        )

    print(
        f"Verification summary: ok={state.ok_count} "
        f"warnings={len(state.warnings)} failures={len(state.failures)}"
    )
    return 1 if state.failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
