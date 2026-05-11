import pandas as pd
import json
import os
import importlib
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from log_utils import log_event


SHEET_TO_GROUP_MAP = {
    # Compute
    "Compute": "Compute",
    "Instance_Pools": "Compute",
    "Instance_Pool_Instances": "Compute",
    "Instance_Configurations": "Compute",
    "Autoscaling_Configurations": "Compute",
    "Autoscaling_Policies": "Compute",
    "Instance_Pool_LB_Attachments": "Compute",
    # Networking
    "Vcn": "Networking",
    "Vcn_Subnets": "Networking",
    "Vcn_Route_Tables": "Networking",
    "Vcn_Route_Rules": "Networking",
    "Vcn_Security_Lists": "Networking",
    "Vcn_Security_Rules": "Networking",
    "Vcn_Network_Security_Groups": "Networking",
    "Vcn_NSG_Rules": "Networking",
    "Vcn_Internet_Gateways": "Networking",
    "Vcn_NAT_Gateways": "Networking",
    "Vcn_Service_Gateways": "Networking",
    "Vcn_Local_Peering_Gateways": "Networking",
    "Vcn_DHCP_Options": "Networking",
    "Vcn_DRG_Attachments": "Networking",
    "Vcn_DRGs": "Networking",
    "Vcn_Virtual_Circuits": "Networking",
    "Subnet": "Networking",
    "Route_Tables": "Networking",
    "Route_Rules": "Networking",
    "Security_Lists": "Networking",
    "Security_Ingress_Rules": "Networking",
    "Security_Egress_Rules": "Networking",
    "Internet_Gateways": "Networking",
    "NAT_Gateways": "Networking",
    "Service_Gateways": "Networking",
    "Service_Gateway_Services": "Networking",
    "Local_Peering_Gateways": "Networking",
    "DHCP_Options": "Networking",
    "Vpn": "Networking",
    "Vpn_Tunnels": "Networking",
    "Fastconnect": "Networking",
    "Fastconnect_Public_Prefixes": "Networking",
    "Fastconnect_Cross_Connect_Mappings": "Networking",
    "Fastconnect_Associated_Tunnels": "Networking",
    "Fastconnect_Bandwidth_Shapes": "Networking",
    "DNS_Zones": "Networking",
    "DNS_Records": "Networking",
    "Load_Balancers": "Networking",
    "LB_Listeners": "Networking",
    "LB_Backend_Sets": "Networking",
    "LB_Backends": "Networking",
    "LB_Hostnames": "Networking",
    "LB_Path_Route_Sets": "Networking",
    "LB_Path_Route_Rules": "Networking",
    "LB_Certificates": "Networking",
    "NLB_Overview": "Networking",
    "NLB_Listeners": "Networking",
    "NLB_Backend_Sets": "Networking",
    "NLB_Backends": "Networking",

    # Storage
    "File_Systems": "Storage",
    "Mount_Targets": "Storage",
    "Export_Sets": "Storage",
    "Exports": "Storage",
    "Snapshots": "Storage",
    "Snapshot_Policies": "Storage",
    "Replications": "Storage",
    "Block_Volumes": "Storage",
    "Boot_Volumes": "Storage",
    "Block_Volume_Attachments": "Storage",
    "Boot_Volume_Attachments": "Storage",
    "Block_Volume_Backups": "Storage",
    "Boot_Volume_Backups": "Storage",
    "Volume_Groups": "Storage",
    "Volume_Group_Backups": "Storage",
    "Object_Storage_Buckets": "Storage",
    "Object_Storage_Retention_Rules": "Storage",

    # Databases (MySQL HeatWave)
    "Mysql": "Databases (MySQL HeatWave)",
    "Mysql_Backups": "Databases (MySQL HeatWave)",

    # Oracle AI Database
    "Dbcs": "Oracle AI Database",
    "Dbcs_Operations": "Oracle AI Database",
    "Dbcs_System_Patches": "Oracle AI Database",
    "Dbcs_DB_Homes": "Oracle AI Database",
    "Dbcs_DB_Home_Patches": "Oracle AI Database",
    "Dbcs_DB_Home_Patch_History": "Oracle AI Database",
    "Dbcs_Databases": "Oracle AI Database",
    "Dbcs_Database_Backups": "Oracle AI Database",
    "Dbcs_Data_Guard": "Oracle AI Database",
    "Dbcs_PDBs": "Oracle AI Database",
    "Dbcs_Nodes": "Oracle AI Database",
    "Adb": "Oracle AI Database",
    "Adb_Backups": "Oracle AI Database",

    # Identity & Security
    "WAF": "Identity & Security",
    "WAF_Policies": "Identity & Security",
    "WAF_Firewalls": "Identity & Security",
    "WAF_Request_Access": "Identity & Security",
    "WAF_Response_Access": "Identity & Security",
    "WAF_Request_Protection": "Identity & Security",
    "WAF_Response_Protection": "Identity & Security",
    "WAF_Request_Rate_Limits": "Identity & Security",
    "WAF_Actions": "Identity & Security",
    "WAF_Edge": "Identity & Security",
    "WAF_Edge_Policies": "Identity & Security",
    "WAF_Edge_Custom_Rules": "Identity & Security",
    "WAF_Edge_Access_Rules": "Identity & Security",
    "WAF_Edge_Protection_Rules": "Identity & Security",
    "WAF_Edge_Rate_Limits": "Identity & Security",
}

GROUP_ORDER = {
    "Compute": 1,
    "Networking": 2,
    "Storage": 3,
    "Oracle AI Database": 4,
    "Databases (MySQL HeatWave)": 5,
    "Identity & Security": 6,
    "Observability and Management": 20,
    "Application Development": 21,
    "Integration": 22,
    "Analytics and AI": 23,
    "Governance and Administration": 24,
    "Migration and Disaster Recovery": 25,
    "Billing and Cost Management": 26,
    "Multicloud": 27,
    "Other": 99,
}


def _get_group_info(sheet_name):
    group_name = SHEET_TO_GROUP_MAP.get(sheet_name, "Other")
    group_no = GROUP_ORDER.get(group_name, 99)
    return group_name, group_no


RAW_COL_NAME_RE = re.compile(r"^[a-z0-9_.]+$")
SHEET_PREFIX_RE = re.compile(r"^(\d+)-")
SHEET_NAME_CANONICAL_MAP = {
    "Compute": "Instance",
    "Instance_Pools": "Instance_Pools",
    "Instance_Pool_Instances": "Instance_Pool_Instances",
    "Instance_Configurations": "Instance_Configurations",
    "Autoscaling_Configurations": "Autoscaling_Configurations",
    "Autoscaling_Policies": "Autoscaling_Policies",
    "Instance_Pool_LB_Attachments": "Instance_Pool_LB_Attachments",
    "Vcn": "VCNs",
    "Vcn_Subnets": "VCN_Subnets",
    "Vcn_Route_Tables": "VCN_Route_Tables",
    "Vcn_Route_Rules": "VCN_Route_Rules",
    "Vcn_Security_Lists": "VCN_Security_Lists",
    "Vcn_Security_Rules": "VCN_Security_Rules",
    "Vcn_Network_Security_Groups": "VCN_Network_Security_Groups",
    "Vcn_NSG_Rules": "VCN_NSG_Rules",
    "Vcn_Internet_Gateways": "VCN_Internet_Gateways",
    "Vcn_NAT_Gateways": "VCN_NAT_Gateways",
    "Vcn_Service_Gateways": "VCN_Service_Gateways",
    "Vcn_Local_Peering_Gateways": "VCN_Local_Peering_Gateways",
    "Vcn_DHCP_Options": "VCN_DHCP_Options",
    "Vcn_DRG_Attachments": "VCN_DRG_Attachments",
    "Vcn_DRGs": "VCN_DRGs",
    "Vcn_Virtual_Circuits": "VCN_Virtual_Circuits",
    "Vpn": "VPN_Connections",
    "Vpn_Tunnels": "VPN_Tunnels",
    "Fastconnect": "FastConnect",
    "Fastconnect_Public_Prefixes": "FastConnect_Public_Prefixes",
    "Fastconnect_Cross_Connect_Mappings": "FastConnect_Cross_Connect_Mappings",
    "Fastconnect_Associated_Tunnels": "FastConnect_Associated_Tunnels",
    "Fastconnect_Bandwidth_Shapes": "FastConnect_Bandwidth_Shapes",
    "DNS_Zones": "DNS_Zones",
    "DNS_Records": "DNS_Records",
    "Dbcs": "DBCS_Systems",
    "Dbcs_Operations": "DBCS_Operations",
    "Dbcs_DB_Homes": "DBCS_DB_Homes",
    "Dbcs_Databases": "DBCS_Databases",
    "Dbcs_Database_Backups": "DBCS_Database_Backups",
    "Dbcs_Data_Guard": "DBCS_Data_Guard",
    "Dbcs_PDBs": "DBCS_PDBs",
    "Dbcs_Nodes": "DBCS_Nodes",
    "Adb": "ADB_Databases",
    "Adb_Backups": "ADB_Backups",
    "Mysql": "MySQL_DB_Systems",
    "Mysql_Backups": "MySQL_Backups",
    "File_Systems": "File_Systems",
    "Mount_Targets": "Mount_Targets",
    "Export_Sets": "Export_Sets",
    "Exports": "Exports",
    "Snapshots": "Snapshots",
    "Snapshot_Policies": "Snapshot_Policies",
    "Replications": "Replications",
    "Load_Balancers": "Load_Balancers",
    "LB_Listeners": "Load_Balancer_Listeners",
    "LB_Backend_Sets": "Load_Balancer_Backend_Sets",
    "LB_Backends": "Load_Balancer_Backends",
    "LB_Hostnames": "Load_Balancer_Hostnames",
    "LB_Path_Route_Sets": "Load_Balancer_Path_Route_Sets",
    "LB_Path_Route_Rules": "Load_Balancer_Path_Routes",
    "LB_Certificates": "Load_Balancer_Certificates",
    "NLB_Overview": "NetworkLoadBalancers",
    "NLB_Listeners": "NetworkLoadBalancerListeners",
    "NLB_Backend_Sets": "NetworkLoadBalancerBackendSet",
    "NLB_Backends": "NetworkLoadBalancerBackends",
    "Block_Volume_Attachments": "Block_Volume_Attachments",
    "Boot_Volume_Attachments": "Boot_Volume_Attachments",
    "Object_Storage_Buckets": "Object_Storage_Buckets",
    "Object_Storage_Retention_Rules": "Object_Storage_Retention_Rules",
    "WAF": "WAF_Policies",
    "WAF_Firewalls": "WAF_Firewalls",
    "WAF_Request_Access": "WAF_Request_Access",
    "WAF_Response_Access": "WAF_Response_Access",
    "WAF_Request_Protection": "WAF_Request_Protection",
    "WAF_Response_Protection": "WAF_Response_Protection",
    "WAF_Request_Rate_Limits": "WAF_Request_Rate_Limits",
    "WAF_Actions": "WAF_Actions",
    "WAF_Edge": "WAF_Edge_Policies",
    "WAF_Edge_Custom_Rules": "WAF_Edge_Custom_Rules",
    "WAF_Edge_Access_Rules": "WAF_Edge_Access_Rules",
    "WAF_Edge_Protection_Rules": "WAF_Edge_Protection_Rules",
    "WAF_Edge_Rate_Limits": "WAF_Edge_Rate_Limits",
}


def _is_raw_col_name(col_name):
    return isinstance(col_name, str) and bool(RAW_COL_NAME_RE.match(col_name))


def _sheet_tab_sort_key(title):
    if title == "Summary":
        return (0, 0, title)
    matched = SHEET_PREFIX_RE.match(str(title))
    if matched:
        return (1, int(matched.group(1)), str(title))
    return (2, 999, str(title))


def _canonical_sheet_name(sheet_name):
    return SHEET_NAME_CANONICAL_MAP.get(sheet_name, sheet_name)


def _category_from_prefixed_sheet_name(prefixed_sheet_name):
    name = str(prefixed_sheet_name)
    if "-" in name:
        return name.split("-", 1)[1]
    return name


def _to_display_column_name(col_name):
    if not isinstance(col_name, str) or "." not in col_name:
        return col_name

    head, tail = col_name.split(".", 1)
    if head.endswith("_raw"):
        head = head[: -len("_raw")]
    elif head.endswith("_enriched"):
        head = head[: -len("_enriched")]
    return f"{head}.{tail}"


def _align_preferred_with_raw(sheet_df, preferred_columns):
    if not preferred_columns:
        return sheet_df, []

    raw_cols = [c for c in sheet_df.columns if _is_raw_col_name(c)]
    resolved_pref = []
    drop_alias_cols = set()

    for pref_col in preferred_columns:
        if pref_col not in sheet_df.columns:
            continue

        if _is_raw_col_name(pref_col):
            if pref_col not in resolved_pref:
                resolved_pref.append(pref_col)
            continue

        alias_series = sheet_df[pref_col]
        matched_raw = None
        for raw_col in raw_cols:
            if raw_col == pref_col:
                continue
            try:
                if sheet_df[raw_col].equals(alias_series):
                    matched_raw = raw_col
                    break
            except Exception:
                continue

        if matched_raw:
            if matched_raw not in resolved_pref:
                resolved_pref.append(matched_raw)
            drop_alias_cols.add(pref_col)
        elif pref_col not in resolved_pref:
            resolved_pref.append(pref_col)

    if drop_alias_cols:
        sheet_df = sheet_df.drop(columns=[c for c in drop_alias_cols if c in sheet_df.columns])

    return sheet_df, resolved_pref


COMMON_CONTEXT_COLUMN_CANDIDATES = [
    (
        ["compartment_name"],
        ["compartment_name", "Compartment", "compartment"],
    ),
    (
        ["region", "region_name"],
        ["region", "region_name", "Region"],
    ),
    (
        ["availability_domain", "availability_domain_name"],
        ["availability_domain", "availability_domain_name", "Availability Domain"],
    ),
    (["fault_domain"], ["fault_domain", "Fault Domain"]),
]


def _pick_service_raw_column(existing, raw_field_names, ordered):
    for suffix in ("_raw", "_enriched"):
        for raw_field_name in raw_field_names:
            for col in existing:
                if col in ordered or not isinstance(col, str) or "." not in col:
                    continue
                head, tail = col.split(".", 1)
                if head.endswith(suffix) and tail == raw_field_name:
                    return col
    return None


def _reorder_common_context_columns(sheet_df):
    if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
        return sheet_df
    if "category" not in sheet_df.columns:
        return sheet_df

    ordered = ["category"]
    existing = list(sheet_df.columns)

    for raw_field_names, fallback_candidates in COMMON_CONTEXT_COLUMN_CANDIDATES:
        selected = _pick_service_raw_column(existing, raw_field_names, ordered)
        if not selected:
            selected = next(
                (c for c in fallback_candidates if c in existing and c not in ordered),
                None,
            )
        if selected:
            ordered.append(selected)

    remaining = [c for c in existing if c not in ordered]
    return sheet_df[ordered + remaining]


def create_report(profile_name, json_paths, tenancy_name=None, extracted_at=None):
    report_dir = "OCI_Reports"
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
        
    output_file = os.path.join(report_dir, f"OCI_Report_{profile_name}.xlsx")
    sheet_count = 0
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # 공통 스타일 정의
    header_core_fill = PatternFill(start_color='A6C9EC', end_color='A6C9EC', fill_type='solid')
    header_detail_fill = PatternFill(start_color='DAE9F8', end_color='DAE9F8', fill_type='solid')
    header_font = Font(bold=True)
    hyperlink_font = Font(color="0000FF", underline="single")
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_side = Side(style='thin')
    double_side = Side(style='double')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(size=10, color="555555")
    meta_label_font = Font(bold=True)
    dashboard_entries = []
    used_sheet_names = set()
    tenancy_value = tenancy_name or profile_name
    extracted_at_value = extracted_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for service_name, path in json_paths.items():
        if not os.path.exists(path):
            log_event(
                "WARN",
                "report",
                "raw_input_missing",
                message="Raw input file is missing. Skipping service sheet generation",
                step_service=service_name,
                path=path,
            )
            continue
        
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if not data:
            log_event(
                "INFO",
                "report",
                "service_data_empty",
                message="No data found for service. Generating empty master sheet",
                step_service=service_name,
            )

        df = pd.json_normalize(data)
        
        # 1. 공통: 카테고리 컬럼 추가
        df.insert(0, 'category', service_name.capitalize())
        
        # 2. 서비스별 전용 포맷터 로드 및 가공
        processed_sheets = {}
        pref_cols_map = {}
        
        try:
            # formatters.formatter_{service_name} 모듈 동적 로드
            module = importlib.import_module(f"formatters.formatter_{service_name}")
            result = module.transform(df)
            
            if isinstance(result, dict):
                processed_sheets = result
            else:
                # 단일 DataFrame 반환 시 기존처럼 서비스명을 시트명으로 사용
                processed_sheets = {service_name.capitalize(): result}
                
            # preferred_columns 가져오기 (dict 또는 list 대응)
            pref = module.get_preferred_columns()
            if isinstance(pref, dict):
                pref_cols_map = pref
            else:
                # 단일 리스트인 경우 첫 번째 시트에 적용
                first_sheet = list(processed_sheets.keys())[0]
                pref_cols_map = {first_sheet: pref}
                
        except ImportError:
            log_event(
                "WARN",
                "report",
                "formatter_not_found",
                message="No specialized formatter found. Using default layout",
                step_service=service_name,
            )
            processed_sheets = {service_name.capitalize(): df}
            pref_cols_map = {}
        except Exception as e:
            log_event(
                "ERROR",
                "report",
                "formatter_error",
                message="Specialized formatter failed. Using default layout",
                step_service=service_name,
                detail=str(e),
            )
            processed_sheets = {service_name.capitalize(): df}
            pref_cols_map = {}

        for sheet_name, sheet_df in processed_sheets.items():
            # 3. 공통: 주요 컬럼을 raw 기준으로 전면 재배치(별칭 중복 제거)
            preferred_columns = pref_cols_map.get(sheet_name, [])
            existing_pref = []

            if preferred_columns:
                sheet_df, resolved_pref = _align_preferred_with_raw(sheet_df, preferred_columns)
                existing_pref = [c for c in resolved_pref if c in sheet_df.columns]
                remaining = [c for c in sheet_df.columns if c not in existing_pref]
                sheet_df = sheet_df[existing_pref + remaining]

            # 4. 엑셀 시트 작성 (대분류 번호 접두 + 표준 시트명 + 31자 제한 + 충돌 회피)
            canonical_sheet_name = _canonical_sheet_name(str(sheet_name))
            group_name, group_no = _get_group_info(str(sheet_name))
            prefixed_sheet_name = f"{group_no}-{canonical_sheet_name}"
            valid_sheet_name = str(prefixed_sheet_name)[:31]
            if valid_sheet_name in used_sheet_names:
                seq = 2
                while True:
                    suffix = f"_{seq}"
                    base = str(prefixed_sheet_name)
                    trimmed = base[: max(1, 31 - len(suffix))]
                    candidate = f"{trimmed}{suffix}"
                    if candidate not in used_sheet_names:
                        valid_sheet_name = candidate
                        break
                    seq += 1

            # category is derived from the final sheet name without numeric prefix.
            category_value = _category_from_prefixed_sheet_name(valid_sheet_name)
            if "category" in sheet_df.columns:
                sheet_df = sheet_df.drop(columns=["category"])
            sheet_df.insert(0, "category", category_value)
            sheet_df = _reorder_common_context_columns(sheet_df)
            preferred_column_names = {"category", *existing_pref}
            preferred_column_positions = [
                idx for idx, col_name in enumerate(sheet_df.columns, 1) if col_name in preferred_column_names
            ]
            # Keep raw schema keys internally, but shorten worksheet headers for readability.
            sheet_df = sheet_df.rename(
                columns={c: _to_display_column_name(c) for c in sheet_df.columns}
            )

            used_sheet_names.add(valid_sheet_name)
            sheet_df.to_excel(writer, sheet_name=valid_sheet_name, index=False)
            dashboard_entries.append(
                {
                    "service": service_name.capitalize(),
                    "group": group_name,
                    "group_no": group_no,
                    "sheet": valid_sheet_name,
                    "rows": len(sheet_df),
                    "columns": len(sheet_df.columns),
                }
            )
            
            # 5. 공통 스타일 및 너비 자동 조정 적용
            ws = writer.sheets[valid_sheet_name]
            total_col_count = ws.max_column
            preferred_column_positions = [
                idx for idx in preferred_column_positions if 1 <= idx <= total_col_count
            ]
            preferred_column_position_set = set(preferred_column_positions)
            last_pref_col_idx = max(preferred_column_positions) if preferred_column_positions else 0

            ws.freeze_panes = "A2"
            if total_col_count > 0 and ws.max_row >= 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(total_col_count)}{ws.max_row}"

            column_widths = {}

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_col_count), 1):
                for col_idx, cell in enumerate(row, 1):
                    cell.border = thin_border
                    
                    if cell.value:
                        length = len(str(cell.value))
                        if col_idx not in column_widths or length > column_widths[col_idx]:
                            column_widths[col_idx] = length
                    
                    if row_idx == 1:
                        cell.fill = (
                            header_core_fill
                            if col_idx in preferred_column_position_set
                            else header_detail_fill
                        )
                        cell.font = header_font
                    
                    if 0 < last_pref_col_idx < total_col_count and col_idx == last_pref_col_idx:
                        cell.border = Border(
                            left=cell.border.left,
                            right=double_side,
                            top=cell.border.top,
                            bottom=cell.border.bottom
                        )
            
            for i, width in column_widths.items():
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = min(width + 4, 50)
            
            sheet_count += 1

    # 6. 요약(목차) 시트 생성: 데이터가 없어도 Summary는 항상 생성한다.
    wb = writer.book
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_dash = wb.create_sheet(title="Summary", index=0)
    writer.sheets["Summary"] = ws_dash

    ws_dash.freeze_panes = "A6"
    ws_dash.merge_cells("A1:E1")
    for col_idx in range(1, 6):
        head_cell = ws_dash.cell(row=1, column=col_idx)
        head_cell.fill = header_core_fill
        head_cell.border = thin_border
    title_cell = ws_dash["A1"]
    title_cell.value = "OCI Resource Report"
    title_cell.font = title_font
    title_cell.alignment = center_align

    ws_dash["D2"] = "테넌시명"
    ws_dash["E2"] = tenancy_value
    ws_dash["D3"] = "추출일자"
    ws_dash["E3"] = extracted_at_value
    for cell_ref in ["D2", "D3"]:
        label_cell = ws_dash[cell_ref]
        label_cell.border = thin_border
        label_cell.fill = header_core_fill
        label_cell.font = meta_label_font
        label_cell.alignment = center_align

    for cell_ref in ["E2", "E3"]:
        value_cell = ws_dash[cell_ref]
        value_cell.border = thin_border
        value_cell.alignment = left_align

    headers = ["그룹 번호", "그룹명", "시트 이름", "리소스 개수", "바로가기"]
    for col_idx, title in enumerate(headers, 1):
        cell = ws_dash.cell(row=5, column=col_idx, value=title)
        cell.fill = header_core_fill
        cell.font = header_font
        cell.border = thin_border

    dashboard_entries_sorted = sorted(
        dashboard_entries,
        key=lambda x: (x["group_no"], x["sheet"]),
    )

    if dashboard_entries_sorted:
        for row_idx, item in enumerate(dashboard_entries_sorted, 6):
            group_no = item["group_no"]
            dashboard_sheet_name = item["sheet"]

            group_no_cell = ws_dash.cell(row=row_idx, column=1, value=group_no)
            group_no_cell.border = thin_border
            group_no_cell.alignment = right_align
            ws_dash.cell(row=row_idx, column=2, value=item["group"]).border = thin_border
            ws_dash.cell(row=row_idx, column=3, value=dashboard_sheet_name).border = thin_border
            rows_cell = ws_dash.cell(row=row_idx, column=4, value=item["rows"])
            rows_cell.border = thin_border
            rows_cell.alignment = right_align

            link_cell = ws_dash.cell(row=row_idx, column=5, value="▶ Go to Sheet")
            link_cell.hyperlink = f"#'{item['sheet']}'!A1"
            link_cell.font = hyperlink_font
            link_cell.alignment = center_align
            link_cell.border = thin_border
    else:
        ws_dash.cell(row=6, column=1, value="-").border = thin_border
        ws_dash.cell(row=6, column=2, value="No Data").border = thin_border
        ws_dash.cell(row=6, column=3, value="-").border = thin_border
        empty_rows_cell = ws_dash.cell(row=6, column=4, value=0)
        empty_rows_cell.border = thin_border
        empty_rows_cell.alignment = right_align
        empty_link_cell = ws_dash.cell(row=6, column=5, value="-")
        empty_link_cell.border = thin_border
        empty_link_cell.alignment = center_align

    if ws_dash.max_row >= 5:
        ws_dash.auto_filter.ref = f"A5:E{ws_dash.max_row}"

    # 대시보드 컬럼 너비
    ws_dash.column_dimensions["A"].width = 10
    ws_dash.column_dimensions["B"].width = 34
    ws_dash.column_dimensions["C"].width = 34
    ws_dash.column_dimensions["D"].width = 12
    ws_dash.column_dimensions["E"].width = 30

    # Keep Summary as the left-most tab and sort service tabs by numeric prefix (ascending).
    desired_order = [ws.title for ws in sorted(wb.worksheets, key=lambda ws: _sheet_tab_sort_key(ws.title))]
    for target_idx, title in enumerate(desired_order):
        ws = next((sheet for sheet in wb.worksheets if sheet.title == title), None)
        if ws is None:
            continue
        current_idx = wb.worksheets.index(ws)
        if current_idx != target_idx:
            wb.move_sheet(ws, offset=target_idx - current_idx)

    writer.close()
    log_event(
        "INFO",
        "report",
        "report_generated",
        message="Excel report generated",
        report=output_file,
        sheet_count=sheet_count + 1,
        resource_sheets=sheet_count,
    )
