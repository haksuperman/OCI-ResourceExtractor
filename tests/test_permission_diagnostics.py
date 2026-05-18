import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import oci
from openpyxl import load_workbook

import runner
from formatters import formatter_base
from log_utils import classify_error, log_event


class FakeServiceError(Exception):
    def __init__(self, status, code="", message=""):
        super().__init__(f"status={status} code={code} message={message}")
        self.status = status
        self.code = code
        self.message = message


class FakeClient:
    profile_name = "unit"
    tenancy_name = "Unit Tenancy"
    tenancy_id = "ocid1.tenancy.oc1..unit"
    config = {}

    def __init__(self, profile_name):
        self.profile_name = profile_name
        self.regions = ["ap-test-1"]
        self.compartments = []


class PermissionDiagnosticsTest(unittest.TestCase):
    def test_error_classification(self):
        cases = [
            (FakeServiceError(401, "NotAuthenticated", "denied"), "permission_denied"),
            (FakeServiceError(403, "Forbidden", "denied"), "permission_denied"),
            (
                FakeServiceError(404, "NotAuthorizedOrNotFound", "denied"),
                "permission_denied",
            ),
            (FakeServiceError(404, "NotFound", "missing"), "not_found"),
            (FakeServiceError(429, "TooManyRequests", "rate"), "rate_limited"),
            (FakeServiceError(500, "InternalError", "service"), "service_error"),
            (Exception("local failure"), "unexpected_error"),
        ]
        for exc, expected in cases:
            with self.subTest(expected=expected):
                self.assertEqual(classify_error(exc), expected)

    def test_log_event_marks_permission_errors_as_warn(self):
        payload = log_event(
            "ERROR",
            "fake",
            "list_failed",
            status=403,
            code="Forbidden",
            message="permission denied",
            notify=False,
        )
        self.assertEqual(payload["level"], "WARN")
        self.assertEqual(payload["kind"], "permission_denied")

    def test_log_event_classifies_oci_service_error_detail_string(self):
        error = oci.exceptions.ServiceError(
            404,
            "NotAuthorizedOrNotFound",
            {},
            "permission denied",
        )
        payload = log_event(
            "ERROR",
            "fake",
            "list_failed",
            detail=str(error),
            notify=False,
        )
        self.assertEqual(payload["level"], "WARN")
        self.assertEqual(payload["kind"], "permission_denied")
        self.assertEqual(payload["status"], "404")
        self.assertEqual(payload["code"], "NotAuthorizedOrNotFound")

    def test_runner_permission_warn_keeps_run_completed(self):
        created_reports = []

        def fake_collector(client):
            path = Path("raw_data") / client.profile_name / f"fake_{client.profile_name}.json"
            path.parent.mkdir(parents=True, exist_ok=True)
            log_event(
                "WARN",
                "fake",
                "fake_listing_failed",
                region=client.regions[0],
                compartment="restricted",
                status=403,
                code="Forbidden",
                message="permission denied",
            )
            path.write_text("[]", encoding="utf-8")
            return str(path)

        def fake_create_report(*args, **kwargs):
            created_reports.append(kwargs)

        raw_payload = None
        with tempfile.TemporaryDirectory() as tmpdir:
            old_cwd = os.getcwd()
            os.chdir(tmpdir)
            try:
                with mock.patch.object(runner.common, "get_profiles", return_value=["unit"]):
                    with mock.patch.object(runner.common, "OCIClient", FakeClient):
                        with mock.patch.object(
                            runner,
                            "SERVICE_REGISTRY",
                            [{"name": "fake", "scope": "regional", "collector": fake_collector}],
                        ):
                            with mock.patch.object(
                                runner.formatter_base,
                                "create_report",
                                side_effect=fake_create_report,
                            ):
                                result = runner.run_inventory(
                                    "unit",
                                    service_names=["fake"],
                                )
                                with open(result["json_paths"]["fake"], "r", encoding="utf-8") as f:
                                    raw_payload = json.load(f)
            finally:
                os.chdir(old_cwd)

        self.assertEqual(result["service_results"][0]["health"], "permission_limited")
        self.assertEqual(result["service_results"][0]["warning_count"], 1)
        self.assertEqual(result["service_results"][0]["permission_denied_count"], 1)
        self.assertEqual(result["service_results"][0]["errors"], 0)
        self.assertEqual(result["run_summary"]["health"], "permission_limited")
        self.assertEqual(result["run_summary"]["permission_denied_count"], 1)
        self.assertTrue(created_reports)
        self.assertEqual(created_reports[0]["diagnostics"][0]["kind"], "permission_denied")
        self.assertEqual(raw_payload, [])

    def test_runner_service_wide_permission_error_writes_empty_raw(self):
        created_reports = []
        raw_payload = None

        def fake_collector(client):
            raise oci.exceptions.ServiceError(
                404,
                "NotAuthorizedOrNotFound",
                {},
                "permission denied",
            )

        def fake_create_report(*args, **kwargs):
            created_reports.append(kwargs)

        with tempfile.TemporaryDirectory() as tmpdir:
            old_cwd = os.getcwd()
            os.chdir(tmpdir)
            try:
                with mock.patch.object(runner.common, "get_profiles", return_value=["unit"]):
                    with mock.patch.object(runner.common, "OCIClient", FakeClient):
                        with mock.patch.object(
                            runner,
                            "SERVICE_REGISTRY",
                            [{"name": "fake", "scope": "regional", "collector": fake_collector}],
                        ):
                            with mock.patch.object(
                                runner.formatter_base,
                                "create_report",
                                side_effect=fake_create_report,
                            ):
                                result = runner.run_inventory(
                                    "unit",
                                    service_names=["fake"],
                                )
                                with open(result["json_paths"]["fake"], "r", encoding="utf-8") as f:
                                    raw_payload = json.load(f)
            finally:
                os.chdir(old_cwd)

        self.assertEqual(raw_payload, [])
        self.assertEqual(result["service_results"][0]["health"], "permission_limited")
        self.assertEqual(result["service_results"][0]["permission_denied_count"], 1)
        self.assertEqual(result["service_results"][0]["errors"], 0)
        self.assertEqual(created_reports[0]["run_summary"]["health"], "permission_limited")

    def test_excel_summary_and_run_diagnostics_sheet(self):
        diagnostics = [
            {
                "timestamp": "2026-05-18 10:00:00",
                "level": "WARN",
                "event": "fake_listing_failed",
                "category": "Run_Diagnostics",
                "health": "permission_limited",
                "service": "fake",
                "region": "ap-test-1",
                "compartment": "restricted",
                "resource": "-",
                "kind": "permission_denied",
                "status": "403",
                "code": "Forbidden",
                "message": "permission denied",
                "detail": "status=403 code=Forbidden message=permission denied",
            }
        ]
        run_summary = {
            "health": "permission_limited",
            "warning_count": 1,
            "permission_denied_count": 1,
            "error_count": 0,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            old_cwd = os.getcwd()
            os.chdir(tmpdir)
            try:
                formatter_base.create_report(
                    "unit",
                    {},
                    tenancy_name="Unit Tenancy",
                    extracted_at="2026-05-18 10:00:00",
                    diagnostics=diagnostics,
                    run_summary=run_summary,
                )
                report_path = Path("OCI_Reports") / "OCI_Report_unit.xlsx"
                wb = load_workbook(report_path, read_only=True, data_only=True)
                self.assertEqual(wb.sheetnames[0], "Summary")
                self.assertIn("99-Run_Diagnostics", wb.sheetnames)
                summary = wb["Summary"]
                self.assertEqual(summary["A1"].value, "OCI Resource Report")
                self.assertEqual(summary["H2"].value, "permission_limited")
                self.assertEqual(summary["H3"].value, 1)
                self.assertEqual(summary["H4"].value, 1)
                self.assertEqual(summary["H5"].value, 0)
                diag = wb["99-Run_Diagnostics"]
                headers = [cell.value for cell in next(diag.iter_rows(min_row=1, max_row=1))]
                self.assertEqual(headers[:11], formatter_base.DIAGNOSTIC_FRONT_COLUMNS)
                first_row = [cell.value for cell in next(diag.iter_rows(min_row=2, max_row=2))]
                self.assertEqual(first_row[0], "Run_Diagnostics")
                self.assertEqual(first_row[6], "permission_denied")
            finally:
                os.chdir(old_cwd)


if __name__ == "__main__":
    unittest.main()
